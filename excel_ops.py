import openpyxl
import random
import time

def excel_oku():
    WorkBook = openpyxl.load_workbook('/Users/ali/Desktop/manken.xlsx')
    sheet = WorkBook['13.03.2020']
    print(sheet["B2"].value)
    row_count = sheet.max_row
    column_count = sheet.max_column

    for i in range(1, row_count + 1):
        for j in range(1, column_count + 1):
            print(sheet.cell(row=i, column=j).value)



def excel_yaz():
    isim_liste = ["Ali", "Aziz", "Ömer", "Can", "Merve", "Emir"]
    soyad_liste = ["Sancar", "Çolak", "Yüret", "Dilmen", "Temiz", "Türkoğlu"]
    WorkBook2 = openpyxl.Workbook()
    sheet = WorkBook2.active
    sheet['A1'] = "No"
    sheet['B1'] = "Ad"
    sheet['C1'] = "Soyad"

    # datayı yazıyoruz

    for i in range(2, 5):
        sheet.cell(row=i, column=1).value = i - 1
        sheet.cell(row=i, column=2).value = random.choice(isim_liste)
        sheet.cell(row=i, column=3).value = random.choice(soyad_liste)

    # Worknook Kaydet
    WorkBook2.save('/Users/ali/Desktop/ogrenci.xlsx')



if __name__ == "__main__":
    excel_yaz()
