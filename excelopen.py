import openpyxl
import random

def excel_oku2():
    WorkBook=openpyxl.load_workbook('/Users/ali/Desktop/manken.xlsx')
    sheet=WorkBook["ogrenciler"]
    ogrno_baslik=sheet['A1']
    ad_baslik=sheet.cell(row=1,column=2)
    print(ogrno_baslik.value)
    print(ad_baslik.value)


    for i in range(1,5):
        print(sheet.cell(row=i, column=1).value,sheet.cell(row=i, column=2).value)

